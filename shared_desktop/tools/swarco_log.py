from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import re
from datetime import datetime, timedelta

def get_fill_color(value):
    if value == '01':
        return PatternFill(start_color='a9d08e', end_color='a9d08e', fill_type='solid')  # зеленый
    elif value == '02':
        return PatternFill(start_color='ffff99', end_color='ffff99', fill_type='solid')  # желтый
    elif value == '04':
        return PatternFill(start_color='ff9999', end_color='ff9999', fill_type='solid')  # красный
    elif value == '06':
        return PatternFill(start_color='ffff99', end_color='ff9999', fill_type='darkUp')  # красно-желтый
    elif value == '10':
        return PatternFill(start_color='ffffff', end_color='a9d08e', fill_type='darkUp')  # зелено-белый
    elif value == 'x':
        return PatternFill(start_color='5DADE2', end_color='5DADE2', fill_type='solid')  # синий
    else:
        return None

def extract_date(line):
    date_pattern = r'\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}'
    match = re.search(date_pattern, line)
    if match:
        return datetime.strptime(match.group(), "%Y-%m-%d %H:%M:%S")
    return None

def get_mode(plan_number):
    if plan_number in ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '14']:
        return "Local"
    elif plan_number == '13':
        return "Sync"
    elif plan_number == '15':
        return "Manual"
    elif plan_number == '16':
        return "Central"
    else:
        return ""

def mask_cs(hex_value):
    # Преобразуем в двоичную строку
    bin_value = bin(int(hex_value, 16))[2:].zfill(8)  # дополняем нулями до 8 бит
    
    # Расшифровка первых 3 бит (КЖЗ)
    kzhz_code = int(bin_value[-3:], 2)
    kzhz_dict = {
        0: "КЖЗ (normal)",
        1: "КЖЗ (manual)",
        2: "КЖЗ (local)",
        3: "Тёмный",
        4: "Желтое мигание",
        5: "Кругом красный",
        6: "Режим ошибки (тёмный)",
        7: "Режим ошибки (желтое мигание)"
    }
    result = kzhz_dict.get(kzhz_code, "Неизвестный режим КЖЗ")
    
    # Расшифровка следующих 2 бит (ручная панель, рабочая станция, форс. режим 'фикстайм')
    panel_station_code = int(bin_value[-5:-3], 2)
    panel_station_dict = {
        1: "ручная панель",
        2: "рабочая станция",
        3: "форс. режим 'фикстайм'"
    }
    if panel_station_code in panel_station_dict:
        result += ", " + panel_station_dict[panel_station_code]
    
    # Расшифровка следующих 2 бит (ручная панель, рабочая станция, форс. глав. контактор вкл)
    contactor_code = int(bin_value[-7:-5], 2)
    contactor_dict = {
        1: "ручная панель",
        2: "рабочая станция",
        3: "форс. глав. контактор вкл"
    }
    if contactor_code in contactor_dict:
        result += ", " + contactor_dict[contactor_code]
    
    # Проверка последнего бита на "главный контактор вкл"
    if int(bin_value[-8], 2) == 1:
        result += ", глав. контактор вкл"
    
    return result

def mask_dll(hex_value, r):
    """Функция для расшифровки состояния детекторов на основе состояния в шестнадцатеричной строке."""
    # Преобразуем шестнадцатеричное значение в десятичное
    try:
        DC = int(hex_value, 16)
    except ValueError:
        print(f"Ошибка преобразования значения '{hex_value}' в шестнадцатеричное число.")
        return []

    occupied_detectors = []

    # Проверяем каждый бит в значении DC
    for i in range(8):
        if DC & (1 << i):  # Если бит равен 1, детектор занят
            detector = i + 1 + 8 * r  # Номер детектора с учетом индексации
            occupied_detectors.append(str(detector))  # Добавляем номер детектора в список
    
    return occupied_detectors

def mask_dl(line):
    """Функция для определения состояния детекторов на основе строки."""

    if len(line) < 33 or line[32] == ' ':
        print("Данные по детекторам отсутствуют.")
        return []

    result = []
    # Проверяем, если 32-й символ это пробел
    if line[31] == ' ':
        return result
    else:
        # Извлекаем шестнадцатеричное значение начиная с 32-го символа до двух пробелов
        hex_value = line[31:].strip()  # Обрезаем строку начиная с 32-го символа и убираем лишние пробелы
        hex_value = hex_value.split()[0]  # Убираем все, что идет после первого пробела

    if len(hex_value) % 2 != 0:
        print(f"Некорректная длина строки: {hex_value}")
        return []

    # Разбиваем строку на блоки по 2 символа
    for r, block in enumerate([hex_value[i:i + 2] for i in range(0, len(hex_value), 2)]):
        occupied_detectors = mask_dll(block, r)
        result.extend(occupied_detectors)

    return result

def process_file(file):
    wb = Workbook()
    ws = wb.active

    header = ['Дата']
    data_rows = []

    lines = file.read().decode('utf-8').splitlines()

    # Обработка строк логов
    for i in range(len(lines)):
        line = lines[i].strip()
        if line:
            date_time = extract_date(line)  # Используем новый метод для извлечения даты
            if date_time:
                if i + 2 < len(lines):
                    state_line = lines[i + 2].strip()
                    if '000' in state_line:
                        parts = state_line.split('000')
                        if len(parts) > 1:
                            data = parts[1].strip()
                            data_digits = ''.join(data.split())
                            groups = [data_digits[j:j + 2] for j in range(0, len(data_digits), 2)]

                            # Извлечение плана и режима
                            plan_line = lines[i + 1].strip()
                            plan_match = re.search(r'\b\d{1,2}\b', plan_line)
                            plan = plan_match.group() if plan_match else ''
                            mode = get_mode(plan)

                            # Извлекаем "Состояние контроллера" (следующие 2 символа после пробела)
                            state_match = re.search(r'\s(\w{2})\s', line)  # Ищем 2 символа после пробела
                            controller_state = state_match.group(1) if state_match else ''  # Берем эти 2 символа

                            # Расшифровка состояния контроллера
                            controller_state_decoded = mask_cs(controller_state)

                            # Вызов функции mask_dl для получения состояния детекторов
                            detector_count = mask_dl(line)

                            row_data = {'timestamp': date_time, 'values': groups, 'plan': plan, 'mode': mode, 'controller_state': controller_state_decoded, 'detector_count': detector_count}
                            data_rows.append(row_data)

    # Заполнение столбцов с учётом времени
    all_times = []
    current_time = data_rows[0]['timestamp'] if data_rows else None

    if current_time:
        last_values = data_rows[0]['values']
        last_plan = data_rows[0]['plan']
        last_mode = data_rows[0]['mode']
        last_controller_state = data_rows[0]['controller_state']
        last_detector_count = data_rows[0]['detector_count']
        for record in data_rows:
            while current_time < record['timestamp']:
                all_times.append({'timestamp': current_time, 'values': last_values, 'plan': last_plan, 'mode': last_mode, 'controller_state': last_controller_state, 'detector_count': last_detector_count})
                current_time += timedelta(seconds=1)
            all_times.append(record)
            current_time = record['timestamp'] + timedelta(seconds=1)
            last_values = record['values']
            last_plan = record['plan']
            last_mode = record['mode']
            last_controller_state = record['controller_state']
            last_detector_count = record['detector_count']

    # Запись заголовков и данных
    max_groups = max(len(row['values']) for row in all_times) if all_times else 0
    header += [f'гр{j+1}' for j in range(max_groups)]
    header.extend(['План', 'Режим', 'Состояние контроллера'])  # Добавляем новый столбец

    # Заголовки для детекторов
    for i in range(128):
        header.append(f'DL{i + 1}')

    ws.append(header)

    # Закрепление первой строки
    ws.freeze_panes = "A2"

    # Применение жирного шрифта к заголовкам
    for cell in ws[1]:  # ws[1] — это первая строка
        cell.font = Font(bold=True)

    for row in all_times:
        row_values = [row['timestamp'].strftime("%Y-%m-%d %H:%M:%S")]
        row_values.extend(row['values'] + [''] * (max_groups - len(row['values'])))
        row_values.append(row['plan'])
        row_values.append(row['mode'])
        row_values.append(row['controller_state'])  # Добавляем расшифрованное состояние контроллера

        # Добавление значений для детекторов
        for i in range(24):
            # Проверяем, занят ли детектор
            if f'{i+1}' in row['detector_count']:
                row_values.append('x')
            else:
                row_values.append('')

        ws.append(row_values)

    for col in ws.columns:
        header_name = col[0].value  # Получаем имя заголовка столбца
        if header_name == 'Дата':
            ws.column_dimensions[col[0].column_letter].width = 20
        elif header_name == 'Режим':
            ws.column_dimensions[col[0].column_letter].width = 8
        elif header_name == 'План':
            ws.column_dimensions[col[0].column_letter].width = 6
        elif header_name == 'Состояние контроллера':
            ws.column_dimensions[col[0].column_letter].width = 55
        else:
            ws.column_dimensions[col[0].column_letter].width = 5

    # Добавление стилей
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            fill_color = get_fill_color(str(cell.value))
            if fill_color:
                cell.fill = fill_color

    return wb


def swarco_log(request):
    if request.method == 'POST' and request.FILES['logfile']:
        file = request.FILES['logfile']

        wb = process_file(file)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="processed_data.xlsx"'
        wb.save(response)
        return response
    
    return render(request, 'tools/swarco_log.html')
